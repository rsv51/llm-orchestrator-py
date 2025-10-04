"""
Excel Import/Export Service

提供提供商、模型配置和模型-提供商关联的批量导入导出功能
使用三工作表架构: Providers, Models, Associations
"""
from typing import List, Dict, Any, Optional, Tuple
from io import BytesIO
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models.provider import Provider, ModelConfig, ModelProvider
from app.core.logger import get_logger

logger = get_logger(__name__)


class ExcelService:
    """Excel 批量导入导出服务 - 三工作表架构"""
    
    def __init__(self, db: AsyncSession):
        self.db = db
    
    # ========================================================================
    # 导出功能
    # ========================================================================
    
    async def export_all(self, include_sample: bool = False) -> BytesIO:
        """
        导出所有配置到单个 Excel 文件(三个工作表)
        
        Args:
            include_sample: 是否包含示例数据
            
        Returns:
            BytesIO: Excel 文件的字节流
        """
        logger.info("Exporting all data to Excel with 3 sheets")
        
        try:
            wb = Workbook()
            
            # 创建三个工作表
            await self._create_providers_sheet(wb, include_sample)
            await self._create_models_sheet(wb, include_sample)
            await self._create_associations_sheet(wb, include_sample)
            
            # 删除默认的Sheet
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            # 保存到字节流
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info("Exported all data successfully")
            return output
            
        except Exception as e:
            logger.error(f"Failed to export all data: {str(e)}", exc_info=True)
            raise
    
    async def _create_providers_sheet(self, wb: Workbook, include_sample: bool = False):
        """创建 Providers 工作表"""
        ws = wb.active
        ws.title = "Providers"
        
        # 设置表头
        headers = ['name', 'type', 'api_key', 'base_url', 'priority', 'weight', 'enabled']
        ws.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 获取现有数据
        query = select(Provider).order_by(Provider.priority.desc(), Provider.id)
        result = await self.db.execute(query)
        providers = result.scalars().all()
        
        for p in providers:
            ws.append([
                p.name,
                p.type,
                p.api_key,
                p.base_url or '',
                p.priority,
                p.weight,
                'true' if p.enabled else 'false'
            ])
        
        # 如果需要示例数据
        if include_sample and not providers:
            ws.append([
                'OpenAI-Main',
                'openai',
                'sk-xxx',
                'https://api.openai.com/v1',
                100,
                100,
                'true'
            ])
            ws.append([
                'Anthropic-Main',
                'anthropic',
                'sk-ant-xxx',
                'https://api.anthropic.com/v1',
                100,
                100,
                'true'
            ])
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
    
    async def _create_models_sheet(self, wb: Workbook, include_sample: bool = False):
        """创建 Models 工作表"""
        ws = wb.create_sheet("Models")
        
        # 设置表头
        headers = ['name', 'remark', 'max_retry', 'timeout']
        ws.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 获取现有数据
        query = select(ModelConfig).order_by(ModelConfig.id)
        result = await self.db.execute(query)
        models = result.scalars().all()
        
        for m in models:
            ws.append([
                m.name,
                m.remark or '',
                m.max_retry,
                m.timeout
            ])
        
        # 如果需要示例数据
        if include_sample and not models:
            ws.append(['gpt-4o', 'GPT-4 Optimized', 3, 60])
            ws.append(['claude-3.5-sonnet', 'Claude 3.5 Sonnet', 3, 60])
        
        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
    
    async def _create_associations_sheet(self, wb: Workbook, include_sample: bool = False):
        """创建 Associations 工作表"""
        ws = wb.create_sheet("Associations")
        
        # 设置表头
        headers = ['model_name', 'provider_name', 'provider_model', 
                   'supports_tools', 'supports_vision', 'weight', 'enabled']
        ws.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 获取现有数据
        query = (
            select(ModelProvider, ModelConfig, Provider)
            .join(ModelConfig, ModelProvider.model_id == ModelConfig.id)
            .join(Provider, ModelProvider.provider_id == Provider.id)
            .order_by(ModelProvider.id)
        )
        result = await self.db.execute(query)
        rows = result.all()
        
        for mapping, model, provider in rows:
            ws.append([
                model.name,
                provider.name,
                mapping.provider_model,
                'true' if mapping.tool_call else 'false',
                'true' if mapping.image else 'false',
                mapping.weight,
                'true' if mapping.enabled else 'false'
            ])
        
        # 如果需要示例数据
        if include_sample and not rows:
            ws.append(['gpt-4o', 'OpenAI-Main', 'gpt-4o-2024-05-13', 'true', 'true', 100, 'true'])
            ws.append(['claude-3.5-sonnet', 'Anthropic-Main', 'claude-3-5-sonnet-20241022', 'true', 'true', 100, 'true'])
        
        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
    
    async def download_template(self, with_sample: bool = False) -> BytesIO:
        """
        下载导入模板
        
        Args:
            with_sample: 是否包含示例数据
            
        Returns:
            BytesIO: Excel 模板文件
        """
        return await self.export_all(include_sample=with_sample)
    
    # ========================================================================
    # 导入功能
    # ========================================================================
    
    async def import_all(self, file: BytesIO) -> Dict[str, Any]:
        """
        从 Excel 导入所有配置(三个工作表)
        
        Args:
            file: Excel 文件的字节流
            
        Returns:
            导入结果统计
        """
        logger.info("Importing all data from Excel with 3 sheets")
        
        try:
            wb = load_workbook(file)
            
            # 创建名称到ID的映射
            provider_map = {}
            model_map = {}
            
            # 导入提供商
            providers_stats = await self._import_providers_sheet(wb, provider_map)
            
            # 导入模型
            models_stats = await self._import_models_sheet(wb, model_map)
            
            # 导入关联
            associations_stats = await self._import_associations_sheet(wb, provider_map, model_map)
            
            # 计算总结
            result = {
                'providers': providers_stats,
                'models': models_stats,
                'associations': associations_stats,
                'summary': {
                    'total_imported': (
                        providers_stats['imported'] + 
                        models_stats['imported'] + 
                        associations_stats['imported']
                    ),
                    'total_skipped': (
                        providers_stats['skipped'] + 
                        models_stats['skipped'] + 
                        associations_stats['skipped']
                    ),
                    'total_errors': (
                        len(providers_stats['errors']) + 
                        len(models_stats['errors']) + 
                        len(associations_stats['errors'])
                    )
                }
            }
            
            logger.info(f"Import completed: {result['summary']}")
            return result
            
        except Exception as e:
            await self.db.rollback()
            logger.error(f"Failed to import data: {str(e)}", exc_info=True)
            raise
    
    async def _import_providers_sheet(self, wb: Workbook, provider_map: Dict[str, int]) -> Dict[str, Any]:
        """导入 Providers 工作表"""
        stats = {
            'total': 0,
            'imported': 0,
            'skipped': 0,
            'errors': []
        }
        
        if 'Providers' not in wb.sheetnames:
            stats['errors'].append({'row': 0, 'field': 'sheet', 'error': 'Providers sheet not found'})
            return stats
        
        ws = wb['Providers']
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        for row_num, row in enumerate(rows, start=2):
            stats['total'] += 1
            
            try:
                if not row or all(cell is None for cell in row):
                    continue
                
                name = str(row[0]).strip() if row[0] else ''
                if not name:
                    stats['errors'].append({
                        'row': row_num,
                        'field': 'name',
                        'error': 'Name is required'
                    })
                    continue
                
                # 检查是否已存在
                query = select(Provider).where(Provider.name == name)
                result = await self.db.execute(query)
                existing = result.scalar_one_or_none()
                
                if existing:
                    provider_map[name] = existing.id
                    stats['skipped'] += 1
                    continue
                
                # 创建提供商
                provider = Provider(
                    name=name,
                    type=str(row[1]).strip() if row[1] else 'openai',
                    api_key=str(row[2]).strip() if row[2] else '',
                    base_url=str(row[3]).strip() if row[3] else None,
                    priority=int(row[4]) if row[4] and str(row[4]).isdigit() else 100,
                    weight=int(row[5]) if row[5] and str(row[5]).isdigit() else 100,
                    enabled=str(row[6]).lower() == 'true' if row[6] else True
                )
                
                self.db.add(provider)
                await self.db.flush()
                
                provider_map[name] = provider.id
                stats['imported'] += 1
                
            except Exception as e:
                stats['errors'].append({
                    'row': row_num,
                    'field': 'database',
                    'error': str(e)
                })
        
        await self.db.commit()
        return stats
    
    async def _import_models_sheet(self, wb: Workbook, model_map: Dict[str, int]) -> Dict[str, Any]:
        """导入 Models 工作表"""
        stats = {
            'total': 0,
            'imported': 0,
            'skipped': 0,
            'errors': []
        }
        
        if 'Models' not in wb.sheetnames:
            stats['errors'].append({'row': 0, 'field': 'sheet', 'error': 'Models sheet not found'})
            return stats
        
        ws = wb['Models']
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        for row_num, row in enumerate(rows, start=2):
            stats['total'] += 1
            
            try:
                if not row or all(cell is None for cell in row):
                    continue
                
                name = str(row[0]).strip() if row[0] else ''
                if not name:
                    stats['errors'].append({
                        'row': row_num,
                        'field': 'name',
                        'error': 'Name is required'
                    })
                    continue
                
                # 检查是否已存在
                query = select(ModelConfig).where(ModelConfig.name == name)
                result = await self.db.execute(query)
                existing = result.scalar_one_or_none()
                
                if existing:
                    model_map[name] = existing.id
                    stats['skipped'] += 1
                    continue
                
                # 创建模型
                model = ModelConfig(
                    name=name,
                    remark=str(row[1]).strip() if row[1] else None,
                    max_retry=int(row[2]) if row[2] and str(row[2]).isdigit() else 3,
                    timeout=int(row[3]) if row[3] and str(row[3]).isdigit() else 30,
                    enabled=True
                )
                
                self.db.add(model)
                await self.db.flush()
                
                model_map[name] = model.id
                stats['imported'] += 1
                
            except Exception as e:
                stats['errors'].append({
                    'row': row_num,
                    'field': 'database',
                    'error': str(e)
                })
        
        await self.db.commit()
        return stats
    
    async def _import_associations_sheet(
        self, 
        wb: Workbook, 
        provider_map: Dict[str, int],
        model_map: Dict[str, int]
    ) -> Dict[str, Any]:
        """导入 Associations 工作表"""
        stats = {
            'total': 0,
            'imported': 0,
            'skipped': 0,
            'errors': []
        }
        
        if 'Associations' not in wb.sheetnames:
            stats['errors'].append({'row': 0, 'field': 'sheet', 'error': 'Associations sheet not found'})
            return stats
        
        ws = wb['Associations']
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        for row_num, row in enumerate(rows, start=2):
            stats['total'] += 1
            
            try:
                if not row or all(cell is None for cell in row):
                    continue
                
                model_name = str(row[0]).strip() if row[0] else ''
                provider_name = str(row[1]).strip() if row[1] else ''
                provider_model = str(row[2]).strip() if row[2] else ''
                
                # 验证必填字段
                if not model_name:
                    stats['errors'].append({
                        'row': row_num,
                        'field': 'model_name',
                        'error': 'Model name is required'
                    })
                    continue
                
                if not provider_name:
                    stats['errors'].append({
                        'row': row_num,
                        'field': 'provider_name',
                        'error': 'Provider name is required'
                    })
                    continue
                
                # 查找模型ID和提供商ID
                model_id = model_map.get(model_name)
                provider_id = provider_map.get(provider_name)
                
                if not model_id:
                    stats['errors'].append({
                        'row': row_num,
                        'field': 'model_name',
                        'error': f"Model '{model_name}' not found"
                    })
                    continue
                
                if not provider_id:
                    stats['errors'].append({
                        'row': row_num,
                        'field': 'provider_name',
                        'error': f"Provider '{provider_name}' not found"
                    })
                    continue
                
                # 检查是否已存在
                query = select(ModelProvider).where(
                    ModelProvider.model_id == model_id,
                    ModelProvider.provider_id == provider_id,
                    ModelProvider.provider_model == provider_model
                )
                result = await self.db.execute(query)
                existing = result.scalar_one_or_none()
                
                if existing:
                    stats['skipped'] += 1
                    continue
                
                # 创建关联
                association = ModelProvider(
                    model_id=model_id,
                    provider_id=provider_id,
                    provider_model=provider_model,
                    tool_call=str(row[3]).lower() == 'true' if row[3] else True,
                    image=str(row[4]).lower() == 'true' if row[4] else False,
                    weight=int(row[5]) if row[5] and str(row[5]).isdigit() else 100,
                    enabled=str(row[6]).lower() == 'true' if row[6] else True
                )
                
                self.db.add(association)
                stats['imported'] += 1
                
            except Exception as e:
                stats['errors'].append({
                    'row': row_num,
                    'field': 'database',
                    'error': str(e)
                })
        
        await self.db.commit()
        return stats